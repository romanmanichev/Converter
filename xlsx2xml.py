import datetime

from openpyxl import load_workbook
from lxml import etree
import datetime as dt
import re
from pandas import ExcelFile


# Функция получения названии листов excel файла
def get_list_from_excel_file(fileEXCEL):
    return ExcelFile(fileEXCEL).sheet_names


# Фильтрация через единый формальный параметр функции (дата, снилс, полис)
def xlsx2xml(fileEXCEL, fileXML, tags, NameOfErrorLogFile, NameOfList='Лист1', FilerOfData=False):
    # Загрузка excel файла и листа
    wb = load_workbook(f"{fileEXCEL}")
    sheet = wb[f"{NameOfList}"]

    # Создание <ZL_LIST> тега
    ZL_LIST = etree.Element("ZL_LIST")

    # Формирование тега <ZGLV>
    ZGLV = etree.SubElement(ZL_LIST, "ZGLV")
    filename = etree.SubElement(ZGLV, "FILENAME")
    data = etree.SubElement(ZGLV, "DATA")
    code_mo = etree.SubElement(ZGLV, "CODE_MO")
    year = etree.SubElement(ZGLV, "YEAR")
    r = etree.SubElement(ZGLV, "R")

    # заполнение тега <ZGLV>
    filename.text = fileXML
    data.text = str(dt.date.today())
    code_mo.text = '352506'
    year.text = str(dt.date.today().year)
    r.text = '1'

    # Счетчик для идентифиактора N_ZAP
    counter = 1

    # Функция для проверки даты по фильтру
    def check_date(olddate, filter):
        try:
            return dt.datetime.strptime(olddate, filter).date()
        except:
            return False

    # Перебор excel файла
    for i in range(1, sheet.max_row + 1):

        # Создание <ZAP> тега
        column = etree.SubElement(ZL_LIST, "ZAP")

        # Счетчик колонки
        colCounter = 1

        # Получение итерируемой строки
        for tag in tags:
            if re.fullmatch(r'\(\D{0,}\)', tag) != None:
                column1 = etree.SubElement(column, tag.strip('(').strip(')'))
                column1.text = ""
            elif tag == "N_ZAP":
                # Добавление идентификатора N_ZAP
                column2 = etree.SubElement(column, "N_ZAP")
                column2.text = str(counter)
                counter += 1
            else:
                column3 = etree.SubElement(column, tag)
                # Условия для проверки некоторых тегов
                otherColumn = str(sheet.cell(row=i, column=colCounter).value)
                if check_date(otherColumn, '%Y-%m-%d %H:%M:%S') != False:
                    column3.text = str(check_date(otherColumn, '%Y-%m-%d %H:%M:%S'))
                elif check_date(otherColumn.strip(' '), '%d.%m.%Y') != False:
                        column3.text = str(check_date(otherColumn.strip(' '), '%d.%m.%Y'))
                else:
                    # Даты, которые не попали под фильтрацию
                    column3.text = otherColumn.strip(' ')

                # Проверка включенного фильтра
                if FilerOfData == True:

                    # Идентификация тега ENP и SNILS
                    if tag == 'ENP':

                        # Проверка тега ENP на длинну 16-ти символов
                        if len(str(sheet.cell(row=i, column=colCounter).value)) != 16:
                            with open(f"{NameOfErrorLogFile}.txt", 'a', encoding='UTF-8') as file:
                                file.write(f"Номер записи {i} Столбец ENP: не содержит 16 цифр\n")
                    elif tag == 'SNILS':
                        # Проверка тега SNLIS на совпадение шаблону
                        match = re.fullmatch(r"\d{3}-\d{3}-\d{3} \d{2}", str(sheet.cell(row=i, column=colCounter).value))
                        if match == None:
                            with open(f"{NameOfErrorLogFile}.txt", 'a', encoding='UTF-8') as file:
                                file.write(f"Номер записи {i} Столбец SNILS: не совпадает шаблону\n")

                colCounter += 1

    # Создание дерева и xml файла
    tree = etree.ElementTree(ZL_LIST)
    tree.write(f"{fileXML}.xml", pretty_print=True, encoding='windows-1251')
