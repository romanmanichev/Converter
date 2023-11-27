from openpyxl import load_workbook
from lxml import etree
import datetime as dt


# Название формируемого xml файла
nameOfFile = 'dp-0101'

# Теги для формарования xml файла
tags = ['N_ZAP', 'FAM', 'IM', 'OT', 'W', 'DR', 'PHONE', 'ENP', 'SNILS'] # 9 


# Загрузка excel файла и листа
wb = load_workbook('listOfDispa.xlsx')
sheet = wb['Лист2']

# Создание <ZL_LIST> тега
ZL_LIST = etree.Element("ZL_LIST")

# Формирование тега <ZGLV>
ZGLV = etree.SubElement(ZL_LIST, "ZGLV")
filename = etree.SubElement(ZGLV, "FILENAME")
data = etree.SubElement(ZGLV, "DATA")
code_mo = etree.SubElement(ZGLV, "CODE_MO")
year = etree.SubElement(ZGLV, "YEAR")
r = etree.SubElement(ZGLV, "R")

# заполнение тега <ZGLV>
filename.text = nameOfFile
data.text = '2023-01-01'
code_mo.text = '352506'
year.text = str(dt.date.today().year)
r.text = '1'

# Счетчик для идентифиактора N_ZAP
counter = 1

# Перебор ecxel файла
for i in range(1, sheet.max_row+1):
	# Создание <ZAP> тега
	column = etree.SubElement(ZL_LIST, "ZAP")
	# Получение итерируемой строки
	for j in range(1, sheet.max_column+1):
	
		# Проврека для N_ZAP и DR
		if tags[j-1] == "N_ZAP":

			# Формирование тега из переменной tags
			column1 = etree.SubElement(column, tags[j-1])

			# Добавление соответствующего текста в тег
			column1.text = str(counter)
			counter += 1
		
		elif tags[j-1] == "DR":

			# Преобразование даты в следующую маску ГГГГ-ММ-ДД
			olddate = str(sheet.cell(row=i, column=j).value)

			column1 = etree.SubElement(column, tags[j-1])

			# Добавление соответствующего текста в тег
			column1.text = olddate

			# Фильтрация даты
			# try:

			# 	newdate = dt.datetime.strptime(olddate, '%Y-%m-%d %H:%M:%S')
			# 	print(newdate.date())

			# except ValueError:

			# 	newdate = dt.datetime.strptime(olddate, '%d.%m.%Y')
			# 	print(newdate.date())
			
			# except:
			# 	print("olddate =", olddate)
				

		else:

			# Формирование тега из переменной tags
			column1 = etree.SubElement(column, tags[j-1])

			# Добавление соответствующего текста в тег
			column1.text = str(sheet.cell(row=i, column=j).value)

	# Ввод в xml файл доп.тегов
	column2 = etree.SubElement(column, "MDR")
	column2.text = str(2)
	column3 = etree.SubElement(column, "ADDRESDP")
	column3.text = str("г.Вологда, ул.Окружное ш., д.3в")
	column4 = etree.SubElement(column, "DISP_TYP")
	column4.text = str(4)
	column5 = etree.SubElement(column, "DATADP")
	column5.text = str("2023-11-20")
	column6 = etree.SubElement(column, "TIMEDP")
	column6.text = str("09:00:00")


# Создание дерева и xml файла 
tree = etree.ElementTree(ZL_LIST)
tree.write(f"{nameOfFile}.xml", pretty_print=True, encoding='windows-1251')
